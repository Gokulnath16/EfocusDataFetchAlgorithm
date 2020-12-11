import  openpyxl as oxl

saleOrderFile = oxl.Workbook()
saleOrderSheet = saleOrderFile.active
saleOrderSheet.cell(row=1, column=1).value = 'Party Name'
saleOrderSheet.cell(row=1, column=2).value = 'Part Number'
saleOrderSheet.cell(row=1, column=3).value = 'Extension'
saleOrderSheet.cell(row=1, column=4).value = 'Quantity'
saleOrderSheet.cell(row=1, column=5).value = 'BatchID'


def createAndWriteInFile(salesOrderDetails, rowCount):
    join = ''
    saleOrderSheet.cell(row=rowCount, column=1).value = salesOrderDetails['partyName']
    saleOrderSheet.cell(row=rowCount, column=2).value = salesOrderDetails['itemName']

    for i in range(len(salesOrderDetails['extension'])):
        get = salesOrderDetails['extension'][i]

        if i == len(salesOrderDetails['extension'])-1:
            join += get
        else:
            join += get + '-'

    saleOrderSheet.cell(row=rowCount, column=3).value = join
    saleOrderSheet.cell(row=rowCount, column=4).value = salesOrderDetails['actualQuantity']
    saleOrderSheet.cell(row=rowCount, column=5).value = salesOrderDetails['batchId']

    saleOrderFile.save('saleOrder.xlsx')

    return

def splitBatchId(batchId):
    batch = list(batchId)
    batch = batch[:len(batch) - 2]
    batch = (''.join(batch))

    return batch


def writeSalesOrderWithDetails(searchRowNum, stockSheetMaxRow, stockSheet, getDayBookDetails, getStockSheetDetails, extension, CountOfRow):
    salesOrderDetails = {}

    if len(searchRowNum) > 1:
        #Get the values using row number(seachRowNum)
        for stockSheetRow in range(len(searchRowNum)):
            getBatchId = stockSheet.cell(row=searchRowNum[stockSheetRow], column=6).value
            getGoodQty = stockSheet.cell(row=searchRowNum[stockSheetRow], column=11).value

            if getGoodQty >= getDayBookDetails['actualQuantity']:
                if splitBatchId(getBatchId) >= '2019':
                    salesOrderDetails['partyName'] = getDayBookDetails['partyName']
                    salesOrderDetails['itemName'] = getDayBookDetails['itemName']
                    salesOrderDetails['actualQuantity'] = getDayBookDetails['actualQuantity']
                    salesOrderDetails['extension'] = extension
                    salesOrderDetails['batchId'] = getBatchId
                    createAndWriteInFile(salesOrderDetails, CountOfRow)
                else:
                    pass
    else:
        # Get the values using row number(seachRowNum)
        for stockSheetRow in range(len(searchRowNum)):
            getBatchId = stockSheet.cell(row=searchRowNum[stockSheetRow], column=6).value
            getGoodQty = stockSheet.cell(row=searchRowNum[stockSheetRow], column=11).value

            if getGoodQty >= getDayBookDetails['actualQuantity']:
                salesOrderDetails['partyName'] = getDayBookDetails['partyName']
                salesOrderDetails['itemName'] = getDayBookDetails['itemCodeWithExtension']
                salesOrderDetails['actualQuantity'] = getDayBookDetails['actualQuantity']
                salesOrderDetails['extension'] = extension
                salesOrderDetails['batchId'] = getBatchId
                createAndWriteInFile(salesOrderDetails, CountOfRow)
            else:
                pass

def splitLast3Digit(itemCode):
    #Split the Last 3 number for search and store the value of last 3 number
    itemcodeWithExtension = itemCode
    splitChar = list(itemCode)
    getExtension = splitChar[-3:]
    getExtension = (''.join((getExtension)))
    getSplitedItemCode = splitChar[: len(splitChar) - 3]
    getSplitedItemCode = (''.join(getSplitedItemCode))

    return  getSplitedItemCode, getExtension, itemcodeWithExtension


def searchForItemNameInStock(getDayBookDetails, stockSheet, dayBookSheetMaxRow, stockSheetMaxRow, CountOfRow):
    getStockSheetDetails = {}
    getSplitItemCodeDetails = {}
    searchRowNum = [] #row1, row2, ....
    extension = []

    #Get StockSheet Details
    for stockSheetRow in range(2, stockSheetMaxRow):
        getItemCode = stockSheet.cell(row=stockSheetRow, column=1).value

        #Split the ItemCode for Search
        getsplitedItemCode = splitLast3Digit(getItemCode)


        #Store the return value from Split
        getSplitItemCodeDetails['itemCode'] = getsplitedItemCode[0]
        getSplitItemCodeDetails['extension'] = getsplitedItemCode[1]
        getSplitItemCodeDetails['itemCodeWithExtension'] = getsplitedItemCode[2]


        #If ItemName of DayBook in ItemCode of Stock Warehouse
        if getSplitItemCodeDetails['itemCode'] == getDayBookDetails['itemName']:
            getDayBookDetails['itemCodeWithExtension'] = getSplitItemCodeDetails['itemCodeWithExtension']
            #Store the row in list
            if getSplitItemCodeDetails['extension'] in extension:
                pass
            else:
                extension.append(getSplitItemCodeDetails['extension'])

            getRowofItemCode = stockSheetRow
            searchRowNum.append(getRowofItemCode)
        else:
            pass

    writeSalesOrderWithDetails(searchRowNum, stockSheetMaxRow, stockSheet, getDayBookDetails, getStockSheetDetails, extension, CountOfRow)

def findItemCodeInDayBook(dayBookSheet, stockSheet, dayBookSheetMaxRow, stockSheetMaxRow, CountOfRow):
    #GET partyName and ItemName and Actual Quantity
    getDayBookDetails = {}

    for dayBookRow in range(2, dayBookSheetMaxRow+1):
        getParytName = dayBookSheet.cell(row=dayBookRow, column=5).value
        getItemName = dayBookSheet.cell(row=dayBookRow, column=10).value
        getActualQuantity = dayBookSheet.cell(row=dayBookRow, column=23).value

        #Assign to Dict
        getDayBookDetails['partyName'] = getParytName
        getDayBookDetails['itemName'] = getItemName
        getDayBookDetails['actualQuantity'] = getActualQuantity

        CountOfRow += 1
        searchForItemNameInStock(getDayBookDetails, stockSheet, dayBookSheetMaxRow, stockSheetMaxRow, CountOfRow)

if __name__ == "__main__":
    CountOfRow = 2

    #Stock WareHouse File
    stock = oxl.load_workbook(filename="Stock.xlsx")
    #DayBook File
    dayBook = oxl.load_workbook(filename="DayBook.xlsx")
    #Activate Files using Active
    stockSheet = stock.active
    dayBookSheet = dayBook.active

    stockSheetMaxRow = stockSheet.max_row
    dayBookSheetMaxRow = dayBookSheet.max_row

    resultFromDaBookForItemCode = findItemCodeInDayBook(dayBookSheet, stockSheet, dayBookSheetMaxRow, stockSheetMaxRow, CountOfRow)
















