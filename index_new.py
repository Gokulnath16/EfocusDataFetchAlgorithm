import  openpyxl as oxl

saleOrderFile = oxl.Workbook()
saleOrderSheet = saleOrderFile.active
saleOrderSheet.cell(row=1, column=1).value = 'Party Name'
saleOrderSheet.cell(row=1, column=2).value = 'Part Number'
saleOrderSheet.cell(row=1, column=3).value = 'Quantity'
saleOrderSheet.cell(row=1, column=4).value = 'BatchID'


class SaleOrder:

    def __init__(self, dayBookSheet, stockSheet, dayBookSheetMaxRow, stockSheetMaxRow, CountOfRow):

        self.stockSheet = stockSheet
        self.dayBookSheet = dayBookSheet
        self.stockSheetMaxRow = stockSheetMaxRow
        self.dayBookSheetMaxRow = dayBookSheetMaxRow
        self.CountOfRow = CountOfRow

# -----------------findItemCodeInDayBook-------------------------------


    def findItemCodeInDayBook(self):
        # GET partyName and ItemName and Actual Quantity
        getDayBookDetailsList = []

        for dayBookRow in range(2, self.dayBookSheetMaxRow + 1):
            getDayBookDetails = {}
            getParytName = self.dayBookSheet.cell(row=dayBookRow, column=5).value
            getItemName = self.dayBookSheet.cell(row=dayBookRow, column=10).value
            getActualQuantity = self.dayBookSheet.cell(row=dayBookRow, column=23).value

            # Assign to Dict
            getDayBookDetails['partyName'] = getParytName
            getDayBookDetails['itemName'] = getItemName
            getDayBookDetails['actualQuantity'] = getActualQuantity

            getDayBookDetailsList.append(getDayBookDetails)

            self.CountOfRow += 1

        #print('getDayBookDetailsList', getDayBookDetailsList)
        return getDayBookDetailsList
            #searchForItemNameInStock(getDayBookDetails, stockSheet, dayBookSheetMaxRow, stockSheetMaxRow, CountOfRow)


# -----------------splitLast3Digit-------------------------------

    def splitLast3Digit(self, itemCode):
        # Split the Last 3 number for search and store the value of last 3 number
        itemcodeWithExtension = itemCode
        splitChar = list(itemCode)
        getExtension = splitChar[-3:]
        getExtension = (''.join((getExtension)))
        getSplitedItemCode = splitChar[: len(splitChar) - 3]
        getSplitedItemCode = (''.join(getSplitedItemCode))

        return getSplitedItemCode, getExtension, itemcodeWithExtension


# -----------------searchForItemNameInStock-------------------------------


    def searchForItemNameInStock(self, getDayBook):
        getStockSheetDetails = {}
        getSplitItemCodeDetails = {}
        searchRowNumOfItems = []
        extension = []
        itemNotFoundInStockSheetList = []
        notFoundIndexList = []

        for getDayBookDetailsRow in range(len(getDayBook)):
            searchRowNum = []  # row1, row2, ....
            rowNumForSearchItemsAvail = {}
            itemNotFound = 0
            searchRowNum.clear()


            # Get StockSheet Details
            for stockSheetRow in range(2, self.stockSheetMaxRow):
                getItemCode = self.stockSheet.cell(row=stockSheetRow, column=1).value

                # Split the ItemCode for Search
                getsplitedItemCode = so.splitLast3Digit(getItemCode)

                # Store the return value from Split
                getSplitItemCodeDetails['itemCode'] = getsplitedItemCode[0]
                getSplitItemCodeDetails['extension'] = getsplitedItemCode[1]
                getSplitItemCodeDetails['itemCodeWithExtension'] = getsplitedItemCode[2]

                # If ItemName of DayBook in ItemCode of Stock Warehouse
                if getSplitItemCodeDetails['itemCode'] == getDayBook[getDayBookDetailsRow]['itemName']:
                    itemNotFound = 1
                    getDayBook[getDayBookDetailsRow]['itemCodeWithExtension'] = getSplitItemCodeDetails['itemCodeWithExtension']
                    # Store the row in list
                    if getSplitItemCodeDetails['extension'] in extension:
                        pass
                    else:
                        extension.append(getSplitItemCodeDetails['extension'])

                    getRowofItemCode = stockSheetRow
                    searchRowNum.append(getRowofItemCode)
                else:
                    pass

            rowNumForSearchItemsAvail['getDayBookDetailsRow'] = getDayBookDetailsRow
            rowNumForSearchItemsAvail['searchRowNum'] = searchRowNum
            
            if itemNotFound != 1:
                notFoundIndexList.append(getDayBookDetailsRow)
                itemNotFoundInStockSheetList.append(getDayBook[getDayBookDetailsRow])


            searchRowNumOfItems.append(rowNumForSearchItemsAvail)

        #writeSalesOrderWithDetails(searchRowNum, stockSheetMaxRow, stockSheet, getDayBookDetails, getStockSheetDetails, extension, CountOfRow)

        return searchRowNumOfItems, getStockSheetDetails, itemNotFoundInStockSheetList, notFoundIndexList

# -----------------splitBatchId-------------------------------

    def splitBatchId(self, batchId):
        batch = list(batchId)
        batch = batch[:len(batch) - 2]
        batch = (''.join(batch))

        return batch

#-----------------writeSalesOrderWithDetails-------------------------------

    def writeSalesOrderWithDetails(self, searchRowNum, getStockSheetDetail, getDayBookDetails):

        salesOrderDetailsList = []
        QtyNotAvailItems = []

        if len(searchRowNum) > 0:

            for getDayBookDetailsRow in range(len(getDayBookDetails)):

                salesOrderDetails = {}
                countQtyofItemList = []
                countQtyofItemList.clear()
                countOfAvailQtyIfBatchQtyNotAvail = getDayBookDetailsRow

                searchRowNumFromSearchRowNum = searchRowNum[getDayBookDetailsRow]['searchRowNum']

                for rowNumCount in range(len(searchRowNumFromSearchRowNum)):
                    countQtyofItem = {}

                    getBatchId = stockSheet.cell(row=searchRowNumFromSearchRowNum[rowNumCount], column=6).value
                    getGoodQty = stockSheet.cell(row=searchRowNumFromSearchRowNum[rowNumCount], column=11).value

                    #print('Avail: ',getGoodQty, '___', 'required' ,getDayBookDetails[getDayBookDetailsRow]['actualQuantity'])

                    if getGoodQty >= getDayBookDetails[getDayBookDetailsRow]['actualQuantity']:
                        countOfAvailQtyIfBatchQtyNotAvail += 1
                        salesOrderDetails['partyName'] = getDayBookDetails[getDayBookDetailsRow]['partyName']
                        salesOrderDetails['itemName'] = getDayBookDetails[getDayBookDetailsRow]['itemCodeWithExtension']
                        salesOrderDetails['actualQuantity'] = getDayBookDetails[getDayBookDetailsRow]['actualQuantity']
                        salesOrderDetails['batchId'] = getBatchId

                        salesOrderDetailsList.append(salesOrderDetails)
                        break

                    else:
                        countQtyofItem['ItemCode'] = getDayBookDetails[getDayBookDetailsRow]['itemCodeWithExtension']
                        countQtyofItem['searchRowNum'] = [searchRowNumFromSearchRowNum[rowNumCount]]
                        countQtyofItem['searchRowNumQty']  = [getGoodQty]
                        countQtyofItem['batchId'] = [getBatchId]

                        countQtyofItemList.append(countQtyofItem)

                if countOfAvailQtyIfBatchQtyNotAvail <= getDayBookDetailsRow:
                    countQtyofItemList[0]['requiredQty'] = getDayBookDetails[getDayBookDetailsRow]['actualQuantity']

                    if sum(countQtyofItemList[0]['searchRowNumQty']) < getDayBookDetails[getDayBookDetailsRow]['actualQuantity']:
                        #message = countQtyofItemList[0]['ItemCode'],  'required ', getDayBookDetails[getDayBookDetailsRow]['actualQuantity'], 'but Actual quantity exist in Stock is ', sum(countQtyofItemList[0]['searchRowNumQty'])
                        mess = '{0} required {1}, but actual quantity exist in stock is {2}'.format(countQtyofItemList[0]['ItemCode'], getDayBookDetails[getDayBookDetailsRow]['actualQuantity'], sum(countQtyofItemList[0]['searchRowNumQty']))
                        QtyNotAvailItems.append(mess)

                    elif sum(countQtyofItemList[0]['searchRowNumQty']) >= getDayBookDetails[getDayBookDetailsRow]['actualQuantity']:
                        getRequiredQty = countQtyofItemList[0]['requiredQty']
                        batchList = []
                        qtyList = []

                        for countQtyRow in range(len(countQtyofItemList[0]['searchRowNumQty'])):
                            getRequiredQty -= countQtyofItemList[0]['searchRowNumQty'][countQtyRow]
                            batchList.append(countQtyofItemList[0]['batchId'][countQtyRow])
                            qtyList.append(countQtyofItemList[0]['searchRowNumQty'][countQtyRow])


                        salesOrderDetails['partyName'] = getDayBookDetails[getDayBookDetailsRow]['partyName']
                        salesOrderDetails['itemName'] = getDayBookDetails[getDayBookDetailsRow]['itemCodeWithExtension']
                        salesOrderDetails['actualQuantity'] = qtyList
                        salesOrderDetails['batchId'] = batchList


                        salesOrderDetailsList.append(salesOrderDetails)


        return salesOrderDetailsList, QtyNotAvailItems


    def saveDataToFile(self, saleOrderData, qtyUnAvail, itemNotFound):
        count = 2

        for Data in range(len(saleOrderData)):
            saleOrderSheet.cell(row=count, column=1).value = saleOrderData[Data]['partyName']
            saleOrderSheet.cell(row=count, column=2).value = saleOrderData[Data]['itemName']
            if type(saleOrderData[Data]['actualQuantity']) is list and type(saleOrderData[Data]['batchId']) is list:
                getQty = saleOrderData[Data]['actualQuantity'].join(',')
                getId =  saleOrderData[Data]['batchId'].join(',')
                saleOrderSheet.cell(row=count, column=4).value = getQty
                saleOrderSheet.cell(row=count, column=5).value = getId
            else:
                saleOrderSheet.cell(row=count, column=4).value = saleOrderData[Data]['actualQuantity']
                saleOrderSheet.cell(row=count, column=5).value = saleOrderData[Data]['batchId']
            count += 1
        saleOrderFile.save('../saleOrder.xlsx')

        return 'Process Completed !', qtyUnAvail, itemNotFound


if __name__ == "__main__":
    CountOfRow = 2

    #Stock WareHouse File
    stock = oxl.load_workbook(filename="../Stock.xlsx")
    #DayBook File
    dayBook = oxl.load_workbook(filename="../DayBook.xlsx")
    #Activate Files using Active
    stockSheet = stock.active
    dayBookSheet = dayBook.active

    stockSheetMaxRow = stockSheet.max_row
    dayBookSheetMaxRow = dayBookSheet.max_row

    so = SaleOrder(dayBookSheet, stockSheet, dayBookSheetMaxRow, stockSheetMaxRow, CountOfRow)


    resultFromDaBookForItemCode = so.findItemCodeInDayBook()
    # print("resultFromDaBookForItemCode",resultFromDaBookForItemCode)

    resultFromItemNameInStock = so.searchForItemNameInStock(resultFromDaBookForItemCode)

    searchRowNum = resultFromItemNameInStock[0]
    itemNotFoundInStockList = resultFromItemNameInStock[2]
    itemNotFoundInStockIndexList = resultFromItemNameInStock[3]

    avaiDayBookDetailInStockList = []
    afterRemoveOfSearchRowNumForNotAvailItemsInStockList = []


    for i in range(len(resultFromDaBookForItemCode)):
        if i in itemNotFoundInStockIndexList:
            pass
        else:
            avaiDayBookDetailInStockList.append(resultFromDaBookForItemCode[i])


    for j in range(len(searchRowNum)):
        if searchRowNum[j]['getDayBookDetailsRow'] in  itemNotFoundInStockIndexList :
            pass
        else:
            afterRemoveOfSearchRowNumForNotAvailItemsInStockList.append(searchRowNum[j])




    resultFromCreateData = so.writeSalesOrderWithDetails(afterRemoveOfSearchRowNumForNotAvailItemsInStockList, resultFromItemNameInStock[1], avaiDayBookDetailInStockList)


    #SaveData to file 
    saveData = so.saveDataToFile(resultFromCreateData[0], resultFromCreateData[1], itemNotFoundInStockList)

    print('Quantity Un Availabel')
    print('')
    for i in range(len(saveData[1])):
        print(saveData[1][i])

    print('')
    print('Item Not Found in Stock List')
    for j in range(len(saveData[2])):
        print(saveData[2][j])

    print('')
    print('')
    print(saveData[0])

